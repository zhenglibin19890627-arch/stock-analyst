"""
US11-EXPORT: 报告导出引擎
========================
生成 Excel (.xlsx) 文件供用户下载。

导出点：
1. 每日报告导出 (daily-report)
2. 自选股总览导出 (watchlist)
3. 回测报告导出 (backtest)

格式要求：
- 表头加粗 + 浅灰背景
- 数字列右对齐
- 评级列按档位着色
- 列宽自适应
- 无数据时写入「暂无数据」提示

依赖：openpyxl>=3.1.0
"""

import io
import logging
from datetime import timedelta, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

_CN_TZ = timezone(timedelta(hours=8), name='Asia/Shanghai')

# ============================================================
# 样式常量
# ============================================================

_HEADER_FONT = Font(bold=True, size=11)
_HEADER_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
_HEADER_ALIGN = Alignment(horizontal='center', vertical='center')
_NUM_ALIGN = Alignment(horizontal='right', vertical='center')
_THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

# 评级着色（需求：强烈推荐=深红，推荐=浅红，持有=灰，减仓=浅绿，卖出=深绿）
_RATING_FILLS = {
    '强烈推荐买入': PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid'),
    '推荐买入': PatternFill(start_color='F1948A', end_color='F1948A', fill_type='solid'),
    '持有观望': PatternFill(start_color='D5D8DC', end_color='D5D8DC', fill_type='solid'),
    '建议减仓': PatternFill(start_color='A9DFBF', end_color='A9DFBF', fill_type='solid'),
    '强烈建议卖出': PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid'),
}
_RATING_FONTS = {
    '强烈推荐买入': Font(color='FFFFFF', bold=True),
    '强烈建议卖出': Font(color='FFFFFF', bold=True),
}


def _style_header(ws, row=1, max_col=None):
    """为表头行应用样式"""
    if max_col is None:
        max_col = ws.max_column
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


def _auto_width(ws):
    """列宽自适应（基于内容长度，中文按2字符宽计算）"""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                val_str = str(cell.value)
                # 中文字符按2倍宽度
                char_len = sum(2 if ord(c) > 127 else 1 for c in val_str)
                max_len = max(max_len, char_len)
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


def _apply_rating_color(ws, col_idx, start_row=2):
    """对评级列应用颜色"""
    for row in range(start_row, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        rating = str(cell.value) if cell.value else ''
        if rating in _RATING_FILLS:
            cell.fill = _RATING_FILLS[rating]
            if rating in _RATING_FONTS:
                cell.font = _RATING_FONTS[rating]


def _write_empty_sheet(ws, msg='暂无数据'):
    """空数据时写入提示"""
    ws.cell(row=1, column=1, value=msg)
    ws.cell(row=1, column=1).font = Font(italic=True, color='999999')


# ============================================================
# 1. 每日报告导出
# ============================================================


def export_daily_report(report_date: str) -> io.BytesIO:
    """生成每日报告 Excel

    Args:
        report_date: 报告日期 YYYY-MM-DD
    Returns:
        BytesIO 缓冲区（可直接作为 Flask send_file 参数）
    """
    from database.db_manager import get_connection

    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT dr.stock_code, dr.stock_name, dr.total_score, dr.rating,
               dr.rating_label, dr.score_change, dr.engine_version,
               dr.key_factors, dr.data_warnings, dr.status, dr.error_msg
        FROM daily_reports dr
        WHERE dr.report_date = ?
        ORDER BY dr.total_score DESC
    """,
        (report_date,),
    )
    rows = [dict(r) for r in cursor.fetchall()]
    conn.close()

    wb = Workbook()

    # --- Sheet1: 概览 ---
    ws1 = wb.active
    ws1.title = '概览'
    headers1 = ['股票名称', '代码', '综合评分', '评级', '较昨日涨跌', '引擎版本']
    ws1.append(headers1)
    _style_header(ws1)

    if not rows:
        _write_empty_sheet(ws1, f'{report_date} 暂无报告数据')
    else:
        for r in rows:
            score_chg = r.get('score_change')
            chg_str = (
                f'+{score_chg}'
                if score_chg and score_chg > 0
                else str(score_chg)
                if score_chg
                else '—'
            )
            ws1.append(
                [
                    r.get('stock_name', ''),
                    r.get('stock_code', ''),
                    r.get('total_score'),
                    r.get('rating', ''),
                    chg_str,
                    r.get('engine_version', ''),
                ]
            )
        # 数字列右对齐
        for row in range(2, ws1.max_row + 1):
            ws1.cell(row=row, column=3).alignment = _NUM_ALIGN
            ws1.cell(row=row, column=5).alignment = _NUM_ALIGN
        _apply_rating_color(ws1, 4)

    # --- Sheet2: 详情 ---
    ws2 = wb.create_sheet('详情')
    headers2 = ['股票名称', '代码', '综合评分', '评级', '操作建议', '数据完整度', '风险提示']
    ws2.append(headers2)
    _style_header(ws2)

    if not rows:
        _write_empty_sheet(ws2, '暂无数据')
    else:
        import json

        for r in rows:
            # 解析 key_factors
            factors_str = ''
            try:
                kf = json.loads(r.get('key_factors') or '{}')
                parts = []
                for dim, info in kf.items():
                    if isinstance(info, dict) and 'score' in info:
                        parts.append(f'{dim}:{info["score"]}分')
                factors_str = ', '.join(parts)
            except (json.JSONDecodeError, TypeError):
                factors_str = str(r.get('key_factors', ''))[:100]

            # 解析 data_warnings
            warnings_str = ''
            try:
                dw = json.loads(r.get('data_warnings') or '[]')
                if isinstance(dw, list):
                    warnings_str = '; '.join(dw[:3])
            except (json.JSONDecodeError, TypeError):
                warnings_str = ''

            ws2.append(
                [
                    r.get('stock_name', ''),
                    r.get('stock_code', ''),
                    r.get('total_score'),
                    r.get('rating', ''),
                    factors_str,
                    '',  # 数据完整度（从 key_factors 提取）
                    warnings_str,
                ]
            )
        _apply_rating_color(ws2, 4)

    _auto_width(ws1)
    _auto_width(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ============================================================
# 2. 自选股总览导出
# ============================================================


def export_watchlist() -> io.BytesIO:
    """生成自选股总览 Excel"""
    from database.db_manager import get_connection

    conn = get_connection()
    cursor = conn.cursor()

    # 获取最新报告日期
    cursor.execute('SELECT MAX(report_date) as d FROM daily_reports')
    row = cursor.fetchone()
    latest_date = row['d'] if row else None

    # 主查询
    if latest_date:
        cursor.execute(
            """
            SELECT s.symbol, s.name, s.market,
                   dr.total_score, dr.rating, dr.rating_label, dr.score_change,
                   h.cost_price, h.quantity,
                   pc.latest_price, pc.pct_change
            FROM stocks s
            LEFT JOIN holdings h ON s.id = h.stock_id
            LEFT JOIN price_cache pc ON s.id = pc.stock_id
            LEFT JOIN daily_reports dr ON s.id = dr.stock_id AND dr.report_date = ?
            WHERE s.status != 'delisted'
            ORDER BY dr.total_score DESC
        """,
            (latest_date,),
        )
    else:
        cursor.execute("""
            SELECT s.symbol, s.name, s.market,
                   NULL as total_score, NULL as rating, NULL as rating_label, NULL as score_change,
                   h.cost_price, h.quantity,
                   pc.latest_price, pc.pct_change
            FROM stocks s
            LEFT JOIN holdings h ON s.id = h.stock_id
            LEFT JOIN price_cache pc ON s.id = pc.stock_id
            WHERE s.status != 'delisted'
            ORDER BY s.added_at DESC
        """)

    stocks = [dict(r) for r in cursor.fetchall()]

    # 资金流向数据
    cursor.execute("""
        SELECT s.symbol, rcf.main_net_inflow, rcf.north_holding_change, rcf.margin_balance
        FROM raw_capital_flow rcf
        JOIN stocks s ON s.id = rcf.stock_id
        WHERE rcf.trade_date = (
            SELECT MAX(trade_date) FROM raw_capital_flow WHERE stock_id = rcf.stock_id
        )
    """)
    capital_map = {}
    for r in cursor.fetchall():
        capital_map[r['symbol']] = dict(r)

    conn.close()

    wb = Workbook()

    # --- Sheet1: 自选股 ---
    ws1 = wb.active
    ws1.title = '自选股'
    headers1 = [
        '股票名称',
        '代码',
        '市场',
        '最新评级',
        '综合评分',
        '今日涨跌%',
        '最新价',
        '持仓成本',
        '持仓数量',
        '盈亏',
    ]
    ws1.append(headers1)
    _style_header(ws1)

    if not stocks:
        _write_empty_sheet(ws1, '暂无自选股数据')
    else:
        for s in stocks:
            market_str = 'A股' if s['market'] == 'a_stock' else '港股'
            latest_price = s.get('latest_price')
            cost = s.get('cost_price')
            qty = s.get('quantity') or 0
            pnl = None
            if latest_price and cost and qty > 0:
                pnl = round((latest_price - cost) * qty, 2)

            ws1.append(
                [
                    s.get('name', ''),
                    s.get('symbol', ''),
                    market_str,
                    s.get('rating', '—'),
                    s.get('total_score'),
                    s.get('pct_change'),
                    latest_price,
                    cost,
                    qty if qty > 0 else None,
                    pnl,
                ]
            )
        # 数字列右对齐
        for row in range(2, ws1.max_row + 1):
            for col in [5, 6, 7, 8, 9, 10]:
                ws1.cell(row=row, column=col).alignment = _NUM_ALIGN
        _apply_rating_color(ws1, 4)

    # --- Sheet2: 资金流向 ---
    ws2 = wb.create_sheet('资金流向')
    headers2 = ['代码', '主力净流入(万)', '北向资金(万)', '融资余额(万)']
    ws2.append(headers2)
    _style_header(ws2)

    has_capital = False
    for s in stocks:
        cap = capital_map.get(s['symbol'], {})
        if cap.get('main_net_inflow') is not None:
            has_capital = True
        ws2.append(
            [
                s.get('symbol', ''),
                cap.get('main_net_inflow'),
                cap.get('north_holding_change'),
                cap.get('margin_balance'),
            ]
        )

    if not has_capital and not stocks:
        _write_empty_sheet(ws2, '暂无资金流向数据')

    for row in range(2, ws2.max_row + 1):
        for col in [2, 3, 4]:
            ws2.cell(row=row, column=col).alignment = _NUM_ALIGN

    _auto_width(ws1)
    _auto_width(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ============================================================
# 3. 回测报告导出
# ============================================================


def export_backtest(market: str = 'a_stock') -> io.BytesIO:
    """生成回测报告 Excel"""
    from database.db_manager import get_connection
    from modules.backtest_engine import BacktestEngine

    engine = BacktestEngine()
    report = engine.compute_market_report(market)

    wb = Workbook()

    # --- Sheet1: 市场报告 ---
    ws1 = wb.active
    ws1.title = '市场报告'

    if report.get('total', 0) == 0:
        _write_empty_sheet(ws1, '暂无回测数据，请先执行批量回测')
    else:
        # 总体指标
        ws1.append(['指标', '数值'])
        _style_header(ws1)
        market_name = 'A股' if market == 'a_stock' else '港股'
        ws1.append(['市场', market_name])
        ws1.append(['回测总数', report.get('total', 0)])
        ws1.append(['总体准确率', f'{report.get("accuracy", 0):.2%}'])
        ws1.append(['正确数', report.get('correct_count', 0)])
        ws1.append(['错误数', report.get('wrong_count', 0)])
        ws1.append(['样本期', report.get('date_range', '')])
        ws1.append([])

        # 分级准确率
        ws1.append(['评级', '总数', '正确', '错误', '准确率', 'T+1平均收益%'])
        _style_header(ws1, row=ws1.max_row, max_col=6)
        rating_stats = report.get('rating_stats', {})
        for rating, stats in rating_stats.items():
            acc = stats.get('accuracy')
            ws1.append(
                [
                    rating,
                    stats.get('total', 0),
                    stats.get('correct', 0),
                    stats.get('wrong', 0),
                    f'{acc:.2%}' if acc is not None else '—',
                    stats.get('avg_return_1d', '—'),
                ]
            )
            _apply_rating_color(ws1, 1, start_row=ws1.max_row)

        ws1.append([])
        # 周期准确率
        ws1.append(['周期', '判定数', '正确', '错误', '准确率', '平均收益%'])
        _style_header(ws1, row=ws1.max_row, max_col=6)
        for period, pdata in report.get('period_accuracy', {}).items():
            pacc = pdata.get('accuracy')
            ws1.append(
                [
                    f'T+{period}',
                    pdata.get('total', 0),
                    pdata.get('correct', 0),
                    pdata.get('wrong', 0),
                    f'{pacc:.2%}' if pacc is not None else '—',
                    pdata.get('avg_return', '—'),
                ]
            )

    # --- Sheet2: 个股明细 ---
    ws2 = wb.create_sheet('个股明细')
    headers2 = [
        '代码',
        '名称',
        '评级日期',
        '评级',
        '评级时价格',
        'T+1收益%',
        'T+5收益%',
        'T+20收益%',
        '判定',
    ]
    ws2.append(headers2)
    _style_header(ws2)

    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT br.rating_date, br.rating, br.price_at_rating,
               br.return_1d, br.return_1w, br.return_1m, br.is_correct,
               s.symbol, s.name
        FROM backtest_results br
        JOIN stocks s ON s.id = br.stock_id
        WHERE br.market = ?
        ORDER BY br.rating_date DESC
        LIMIT 200
    """,
        (market,),
    )
    detail_rows = [dict(r) for r in cursor.fetchall()]
    conn.close()

    if not detail_rows:
        _write_empty_sheet(ws2, '暂无个股回测明细')
    else:
        for r in detail_rows:
            correct_map = {1: '✓正确', 0: '✗错误', None: '—中性'}
            ws2.append(
                [
                    r.get('symbol', ''),
                    r.get('name', ''),
                    r.get('rating_date', ''),
                    r.get('rating', ''),
                    r.get('price_at_rating'),
                    r.get('return_1d'),
                    r.get('return_1w'),
                    r.get('return_1m'),
                    correct_map.get(r.get('is_correct'), '—'),
                ]
            )
        for row in range(2, ws2.max_row + 1):
            for col in [5, 6, 7, 8]:
                ws2.cell(row=row, column=col).alignment = _NUM_ALIGN
        _apply_rating_color(ws2, 4)

    _auto_width(ws1)
    _auto_width(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
